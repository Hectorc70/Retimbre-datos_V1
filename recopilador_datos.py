#import pdb; pdb.set_trace()
"""EXCEL"""
import openpyxl
from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.styles.colors import YELLOW


"""INTERFAZ"""
import tkinter as tk
from tkinter.filedialog import*
from tkinter import ttk
from tkinter import filedialog
from tkinter import scrolledtext
from tkinter import ttk
from PIL import ImageTk,Image

#rutas
import sys
import os









class Carpetas:

	"""Recopila los datos necesarios para el proceso del
	   Retimbrado de los empleados de BASE deñ H Ayuntamiento"""
	def __init__ (self):
		self.ventana_prin 		= tk.Tk()											   #inicia la ventana
		self.tam	       		= self.ventana_prin.geometry("800x600+200+50")		   #indica dimenciones y posicion de la ventanna
		self.redimension  		= self.ventana_prin.resizable(width=False, height=False) #indica si la ventana se puede redimencionar
		#self.ventana_color		= self.ventana_prin.configure(background="#F5EEEE")
		self.salida_texto       = scrolledtext.ScrolledText(self.ventana_prin,width=90,height=15)
		self.barra_progreso 	= ttk.Progressbar()
		#self.icono 				= self.ventana_prin.iconbitmap('ico_p.ico')
		self.nombre_v           = self.ventana_prin.title("Recopilacion de Datos - RECALCULO")



		self.iq_ccn_conte         = list()
		self.documentos_dicc      = {}
		self.rutas_subcarpetas    = list()
		self.titulos_completos_iq = list()

		self.no_control = list()
		self.ccn1401 = list()
		self.ccn2240 = list()
		self.ccn2566 = list()
		self.ccn481 = list()
		self.ccn559 = list()

		#self.path_excel			= filedialog.askopenfilename()

		#self.wb          			= openpyxl.Workbook()

		#self.titulo_hoja_activa		= self.wb.active.title

		self.titulo_fila_control, self.titulo_columna_control = 1,1
		self.titulo_fila_path, self.titulo_columna_path	 	  = 1,2

		self.control_fila_txt1, self.control_columna_txt1 = 1,1
		self.path_fila_txt1, self.path_columna_txt1		  = 1,2

		self.control_fila_txt4, self.control_columna_txt4 = 1,1
		self.path_fila_txt4, self.path_columna_txt4		  = 1,2

		self.control_fila_xml1, self.control_columna_xml1 = 1,1
		self.path_fila_xml1, self.path_columna_xml1		  = 1,2

		self.control_fila_xml4, self.control_columna_xml4 = 1,1
		self.path_fila_xml4, self.path_columna_xml4		  = 1,2

		self.fila_xml1_uuid, self.columna_xml1_uuid = 1,3
		self.fila_xml1_arch, self.columna_xml1_arch = 1,4

		self.fila_xml4_uuid, self.columna_xml14_uuid = 1,3
		self.fila_xml4_arch, self.columna_xml4_arch = 1,4

		self.fila_control_sap, self.columna_control_sap = 1,1
		self.fila_nom1_sap, self.columna_nom1_sap = 1,2
		self.fila_nom4_sap, self.columna_nom4_sap = 1,3

		self.ccn1401_fila_iq, self.ccn2240_fila_iq = 1,1
		self.ccn2566_fila_iq, self.ccn559_fila_iq = 1,1
		self.control_fila_iq, self.ccn481_fila_iq=1,1


	def rutas_docs(self):

		"""Recupera y formatea las rutas de los documentos
		   que contienen los datos de entrada """

		self.docs_trabajo       = filedialog.askdirectory()
		self.lista_docs_trabajo = os.listdir(self.docs_trabajo)

		self.imprimir           = """---Se Cargaron los documentos---:
										'\n'{}'\n'{}'\n'{}'\n'{}'\n'{}'\n'""".format(str(self.lista_docs_trabajo[0]),
																					 str(self.lista_docs_trabajo[1]),
																					 str(self.lista_docs_trabajo[2]),
																					 str(self.lista_docs_trabajo[3]),
																					 str(self.lista_docs_trabajo[4]))







		for documentos in self.lista_docs_trabajo:

			documentos_div = documentos.split("_")
			ruta_carpeta   = self.docs_trabajo.replace("/","\\")

			if documentos_div[1] != "modelo(retimbre).xlsx" and documentos_div[0] =="IQ":

				self.documentos_dicc["IQ"] = ruta_carpeta + "\\" + documentos

			if documentos_div[1] == "modelo(retimbre).xlsx":
				self.documentos_dicc["IQ_TRABAJO"] = ruta_carpeta + "\\" + documentos

			if documentos_div[0] == "ReporteSAP" and documentos_div[1] == "Base":
				self.documentos_dicc["REPORTE_SAP_BASE"] = ruta_carpeta + "\\" + documentos

			if documentos_div[0] == "ReporteTimbrado" and documentos_div[1] == "BASE":
				self.documentos_dicc["REPORTE_TIM_1"] = ruta_carpeta + "\\" + documentos

			if documentos_div[0] == "ReporteTimbrado" and documentos_div[1] == "BASE4":
				self.documentos_dicc["REPORTE_TIM_4"] = ruta_carpeta + "\\" + documentos
			else:

				self.documentos_dicc["otro"] = ruta_carpeta + "\\" + documentos

		self.salida_texto.insert(INSERT, self.imprimir)
		self.barra_progreso.step(10)


	def lectura_de_carpeta(self):
		""" lee las carpetas donde se almacenan los achivos
			txt y xml """
		self.ruta_carpeta   		 = filedialog.askdirectory()
		self.lista_carpetas_leidas 	 = os.listdir(self.ruta_carpeta)


		""" Recorre la lista de las subcarpetas"""
		for self.subcarpetas in self.lista_carpetas_leidas:


			ruta_completa_subcarpeta 			= self.ruta_carpeta +"\\" + self.subcarpetas + "\\"
			ruta_completa_subcarpeta_correcta   = ruta_completa_subcarpeta.replace("/","\\")
			self.rutas_subcarpetas.append(ruta_completa_subcarpeta_correcta)




			"""if disco_de_ruta !='C':
				ruta_completa_correcta_disco = ruta_completa_correcta.replace(disco_de_ruta,"C:")
				#print(ruta_completa_correcta_disco)

				self.ruta_carpeta_final = ruta_completa_correcta_disco

				self.rutas.append(self.ruta_carpeta_final)"""

		self.salida_texto.insert(INSERT, "----Lectura de Directorios Terminada----")
		self.barra_progreso.step(20)


	def lectura_de_archivos(self):
		"""Recupera, guarda los archivos XML y TXT
		   clasificandolos por tipo de empleado y archivo"""


		self.ruta_completa_archivos_list_xml1 = list()
		self.ruta_completa_archivos_list_txt1 = list()

		self.ruta_completa_archivos_list_xml4 = list()
		self.ruta_completa_archivos_list_txt4 = list()

		self.ruta_completa_archivos_list_xml5 = list()
		self.ruta_completa_archivos_list_txt5 = list()




		for self.rutas_carpetas in self.rutas_subcarpetas:

			self.archivos  = os.listdir(self.rutas_carpetas)

			for archivos in self.archivos:

				self.ruta_completa_archivos_Z = self.rutas_carpetas + archivos
				self.ruta_completa_archivos = self.ruta_completa_archivos_Z.replace("Z","C")
				self.ruta_completa_archivos.split("\\")

				self.ruta_completa_archivos_div = self.ruta_completa_archivos.split("\\")
				#print(self.ruta_completa_archivos_div)
				tipo_de_nom_carpeta 			= self.ruta_completa_archivos_div[-2]
				self.periodo 						= self.ruta_completa_archivos_div[2]

				tipo_de_nom_carpeta.split("_")
				tipo_de_nom_carpeta_div 		= tipo_de_nom_carpeta.split("_")
				tipo_de_nomina 					= tipo_de_nom_carpeta_div[0]




				self.nombre_de_archivo = self.ruta_completa_archivos_div[-1]
				self.nombre_de_archivo.split(".")
				self.extencion_de_archivo = self.nombre_de_archivo.split(".")

				tipo_de_archivo = self.extencion_de_archivo[-1]
				#print(tipo_de_archivo)

				nombre_del_archivo_solo = self.extencion_de_archivo[0]
				nombre_del_archivo_solo.split("_")
				nombre_logs = nombre_del_archivo_solo.split("_")
				#print(nombre_logs)


				if nombre_logs[0] != "Log" and tipo_de_nomina == "BASE" or tipo_de_nomina =="base":
					if tipo_de_archivo == "xml" or tipo_de_archivo =="XML" :
						self.ruta_completa_archivos_list_xml1.append(self.ruta_completa_archivos)
					if tipo_de_archivo == "txt" or tipo_de_archivo == "TXT":
						self.ruta_completa_archivos_list_txt1.append(self.ruta_completa_archivos)


				if nombre_logs[0] != "Log" and tipo_de_nomina == "BASE4" or tipo_de_nomina == "base4":
					if tipo_de_archivo == "xml" or tipo_de_archivo == "XML":
						self.ruta_completa_archivos_list_xml4.append(self.ruta_completa_archivos)
					if tipo_de_archivo == "txt" or tipo_de_archivo == "TXT":
						self.ruta_completa_archivos_list_txt4.append(self.ruta_completa_archivos)


				if nombre_logs[0] != "Log" and tipo_de_nomina == "CONFIANZA" or tipo_de_nomina == "confianza":
					if tipo_de_archivo == "xml" or tipo_de_archivo == "XML":
						self.ruta_completa_archivos_list_xml5.append(self.ruta_completa_archivos)
					if tipo_de_archivo == "txt" or tipo_de_archivo == "TXT":
						self.ruta_completa_archivos_list_txt5.append(self.ruta_completa_archivos)

		self.barra_progreso.step(30)


	def control_de_listados(self):
		"""Recupera y guarda los No de control
		   de los archivos"""

		self.lectura_de_archivos()

		self.control_list_xml1 = list()
		self.control_list_txt1 = list()

		self.control_list_xml4 = list()
		self.control_list_txt4 = list()

		self.control_list_xml5 = list()
		self.control_list_txt5 = list()



		for rutas_archivos_xml1 in self.ruta_completa_archivos_list_xml1:

			rutas_archivos_xml1.split("\\")
			rutas_archivos_divididas = rutas_archivos_xml1.split("\\")
			nombre_de_archivo = rutas_archivos_divididas[-1]


			nombre_de_archivo.split("_")
			nombre_de_archivo_div = nombre_de_archivo.split("_")
			control = nombre_de_archivo_div[0]

			self.control_list_xml1.append(control)


		for rutas_archivos_txt1 in self.ruta_completa_archivos_list_txt1:

			rutas_archivos_txt1.split("\\")
			rutas_archivos_divididas = rutas_archivos_txt1.split("\\")
			nombre_de_archivo = rutas_archivos_divididas[-1]


			nombre_de_archivo.split("_")
			nombre_de_archivo_div = nombre_de_archivo.split("_")
			control = nombre_de_archivo_div[0]

			self.control_list_txt1.append(control)



		for rutas_archivos_xml4 in self.ruta_completa_archivos_list_xml4:

			rutas_archivos_xml4.split("\\")
			rutas_archivos_divididas = rutas_archivos_xml4.split("\\")



			#print(rutas_archivos_divididas)
			nombre_de_archivo = rutas_archivos_divididas[-1]


			nombre_de_archivo.split("_")
			nombre_de_archivo_div = nombre_de_archivo.split("_")
			control = nombre_de_archivo_div[0]

			self.control_list_xml4.append(control)


		for rutas_archivos_txt4 in self.ruta_completa_archivos_list_txt4:

			rutas_archivos_txt4.split("\\")
			rutas_archivos_divididas = rutas_archivos_txt4.split("\\")
			nombre_de_archivo = rutas_archivos_divididas[-1]

			nombre_de_archivo.split("_")
			nombre_de_archivo_div = nombre_de_archivo.split("_")
			control = nombre_de_archivo_div[0]

			self.control_list_txt4.append(control)



		self.barra_progreso.step(40)


	def lectura_de_UUID(self):
		"""Lee los resportes de del timbrado de BASE.
			Recupera y almacena los UUID de cada empleado"""

		self.ruta_excel_report_b1 = self.documentos_dicc["REPORTE_TIM_1"]
		self.ruta_excel_report_b4 = self.documentos_dicc["REPORTE_TIM_4"]




		self.ruta_report_b1 		= openpyxl.load_workbook(self.ruta_excel_report_b1)
		self.titulos 				= list()
		self.hoja_lista				= list()
		self.archivo_xml1_lista		= list()
		self.uuid_xml1_lista		= list()

		self.ruta_report_b4 		= openpyxl.load_workbook(self.ruta_excel_report_b4)
		self.archivo_xml4_lista		= list()
		self.uuid_xml4_lista		= list()
		self.titulos4 				= list()
		self.hoja_lista_n4			= list()



		"""            NOM1              """

		for self.hojas in self.ruta_report_b1.worksheets:
			self.hoja_lista.append(self.hojas)


		hoja1 = self.hoja_lista[0]
		celdas_titulos = hoja1['A1':'AD1']


		for fila in celdas_titulos:
			for celda in fila:
				self.titulos.append(celda.value)


				if celda.value == "ARCHIVO":
					columnas_archivo = hoja1["A2" : "A1680"]

					for column_archivo in columnas_archivo:
						for columna_celda_archivo in column_archivo:
							#print(columna_celda_archivo.value)
							self.archivo_xml1_lista.append(columna_celda_archivo.value)





				if celda.value == "UUID":
					celda_uuid = (celda.value)
					columnas_uuid  = hoja1["I2" : "I1680"]

					for columnauuid in columnas_uuid:
						for columna_celda in columnauuid:
							self.uuid_xml1_lista.append(columna_celda.value)
							#print(columna_celda.value)




		for self.hojas_n4 in self.ruta_report_b4.worksheets:
			self.hoja_lista_n4.append(self.hojas_n4)

		hoja1_n4 = self.hoja_lista_n4[0]
		celdas_titulos_n4 = hoja1_n4['A1':'AD1']


		for fila4 in celdas_titulos_n4:
			for celda4 in fila4:
				self.titulos4.append(celda4.value)


				if celda4.value == "ARCHIVO":
					columnas_archivo_n4 = hoja1_n4["A2" : "A1680"]

					for column_archivo_n4 in columnas_archivo_n4:
						for columna_celda_archivo_n4 in column_archivo_n4:
							#print(columna_celda_archivo_n4.value)
							self.archivo_xml4_lista.append(columna_celda_archivo_n4.value)





				if celda4.value == "UUID":
					celda_uuid_n4 = (celda4.value)
					columnas_uuid_n4  = hoja1_n4["I2" : "I1680"]

					for columnauuid_n4 in columnas_uuid_n4:
						for columna_celda_n4 in columnauuid_n4:

							self.uuid_xml4_lista.append(columna_celda_n4.value)
							#print(columna_celda_n4.value)


		self.ruta_report_b1.close()
		self.ruta_report_b4.close()

		self.barra_progreso.step(50)


	def lectura_reporte_SAP(self):
		"""Lee el Reporte de los TXT bajado de SAP"""


		self.hoja_lista = list()
		self.titulos 	= list()
		self.control    = list()
		self.nom1    	= list()
		self.nom4    	= list()

		self.ruta_reporte_sap = self.documentos_dicc["REPORTE_SAP_BASE"]
		self.ruta_doc = load_workbook(self.ruta_reporte_sap)



		for hojas in self.ruta_doc.worksheets:
			self.hoja_lista.append(hojas)


		hoja1 = self.hoja_lista[0]
		celdas_titulos = hoja1['A1':'C1']
		columna_max = hoja1.max_column
		fila_max 	= hoja1.max_row

		for fila in celdas_titulos:
			for celda in fila:
				self.titulos.append(celda.value)



				if celda.value == "NOCONTROL" or  celda.value == "NO CONTROL" or celda.value == "CONTROL":

					celdas_control = hoja1['A2':'A'+str(fila_max)]

					for fila_control in celdas_control:
						for celda_control in fila_control:
							self.control.append(celda_control.value)


				if celda.value == "NOM1" or celda.value == "ORDINARIA":
					celdas_nom1 = hoja1['B2':'B'+str(fila_max)]

					for fila_nom1 in celdas_nom1:
						for celda_nom1 in fila_nom1:
							self.nom1.append(celda_nom1.value)


				if celda.value == "NOM4" or celda.value == "ASS":
					celdas_nom4 = hoja1['C2':'C'+str(fila_max)]

					for fila_nom4 in celdas_nom4:
						for celda_nom4 in fila_nom4:
							self.nom4.append(celda_nom4.value)
		#print(self.control)
		"""print(self.nom1)
								print(self.nom1)"""
		self.ruta_doc.close()
		self.barra_progreso.step(60)


	def lectura_IQ(self):
		"""Lee el Archivo del IQ y almacena la info
			"""


		self.ruta_iq			    = self.documentos_dicc["IQ"]
		self.doc_iq 				= openpyxl.load_workbook(self.ruta_iq)
		self.hoja_lista 		 	= list()

		self.empleados_dicc			= {}
		self.empleados_list			= list()

		self.celdas_dicc_ccn	 	= {}
		self.celdas_dicc_nombres 	= {}
		self.claves_ccn_inicio		= {}
		self.claves_ccn_fin	    	= {}
		self.claves_nombres_inicio	= {}
		self.claves_nombres_fin		= {}

		self.conte_iq				 = list()
		self.dicc_control			 = {}

		self.celdas_list_ccn	 = list()
		self.celdas_list_nombres = list()


		self.area_personal  = list()
		self.ccn793 = list()
		self.ccn794 = list()
		self.ccn1000 = list()
		self.ccn1101 = list()
		self.ccn1102 = list()
		self.ccn1104 = list()
		self.ccn1112 = list()
		self.ccn1114 = list()
		self.ccn1115 = list()
		self.ccn1116 = list()
		self.ccn1119 = list()
		self.ccn1129 = list()
		self.ccn1133 = list()
		self.ccn1136 = list()
		self.ccn1148 = list()
		self.ccn1159 = list()

		self.ccn1415 = list()
		self.ccn1501 = list()
		self.ccn1502 = list()
		self.ccn1512 = list()
		self.ccn1536 = list()
		self.ccn1610 = list()
		self.ccn1620 = list()
		self.ccn1645 = list()
		self.ccn2140 = list()
		self.ccn2141 = list()
		self.ccn2144 = list()
		self.ccn2145 = list()
		self.ccn2148 = list()
		self.ccn2150 = list()
		self.ccn2151 = list()
		self.ccn2155 = list()
		self.ccn2200 = list()

		self.ccn2450 = list()
		self.ccn2550 = list()

		self.ccn2587 = list()
		self.ccn2596 = list()
		self.ccn2597 = list()
		self.ccn2598 = list()
		self.ccn2691 = list()
		self.ccn300 = list()
		self.ccn305 = list()
		self.ccn306 = list()
		self.ccn380 = list()
		self.ccn391 = list()

		self.ccn494 = list()










		for self.hojas in self.doc_iq.worksheets:
			self.hoja_lista.append(self.hojas)



		hoja1 = self.hoja_lista[0]
		columna_max = hoja1.max_column
		fila_max 	= hoja1.max_row


		titulos_ccn_c= hoja1["A2":"CU2"]
		titulos_nombres_c = hoja1["A3":"CU3"]

		celdas_titulos = hoja1["A2":"CU2"]
		celdas_titulos_nombres = hoja1["A3":"CU3"]




		"""NOMBRES_TITULOS_CCN"""
		for fila in celdas_titulos:
			for celda in fila:
				if celda.value is not None:
					self.celdas_dicc_ccn[celda.value] = celda


		"""NOMBRES_TITULOS"""
		for fila in celdas_titulos_nombres:
			for celda in fila:
				if celda.value is not None:
					self.celdas_dicc_nombres[celda.value] = celda


		"""CONCATENA LAS DOS FILAS DE TITULOS"""

		for fila, fila_ccn in zip(titulos_nombres_c,titulos_ccn_c):
			for celda, celda_ccn in zip(fila, fila_ccn):
				self.celdas_list_ccn.append(celda_ccn.value)
				self.celdas_list_nombres.append(celda.value)

		for titulos_ccn, titulos_nombres  in zip(self.celdas_list_ccn,self.celdas_list_nombres):

			titulos_completos = str(titulos_ccn) + " " + str(titulos_nombres)
			self.titulos_completos_iq.append(titulos_completos)





		"""OBTENER MINIMO Y MAXIMO DE LAS COLUMNAS"""


		for (celda_nombres_text,celda_nombres_clv) in self.celdas_dicc_nombres.items():
			self.nombres_celd_div = str(celda_nombres_clv).split(".")
			self.columna 	  	  =	self.nombres_celd_div[1][:-2]       				 #Empieza el rango en la linea 3
			self.columna_num  	  =	self.nombres_celd_div[1][-2:-1]
			self.columna_sum      =	int(self.columna_num) + 1
			self.columna_inicio = str(self.columna) + str(self.columna_sum)
			self.columna_fin    = str(self.columna) + str(fila_max)
			self.claves_nombres_inicio[celda_nombres_text] = self.columna_inicio
			self.claves_nombres_fin[celda_nombres_text]    = self.columna_fin

		for celda_ccn_text,celda_ccn_clav in self.celdas_dicc_ccn.items():
			self.ccn_celd_div 	= str(celda_ccn_clav).split(".")
			self.columna_ccn 	  			  	      =	self.ccn_celd_div[1][:-2]
			self.columna_num_ccn  					  =	self.ccn_celd_div[1][-2:-1]
			self.columna_sum_ccn  			 	      =	int(self.columna_num_ccn) + 2
			self.columna_inicio_ccn 				  = str(self.columna_ccn) + str(self.columna_sum_ccn)
			self.columna_fin_ccn  					  = str(self.columna_ccn) + str(fila_max)
			self.claves_ccn_inicio[celda_ccn_text] 	  = self.columna_inicio_ccn
			self.claves_ccn_fin[celda_ccn_text] 	  = self.columna_fin_ccn







		"""CONTENIDO DE IQ COLUMNAS"""

		for titulos_celdas,titulos_celdas_nombres in zip(
														self.celdas_dicc_ccn.keys(),
														self.celdas_dicc_nombres.keys()
														):

			celdas_nombres  = hoja1[str(self.claves_nombres_inicio[titulos_celdas_nombres]):
									str(self.claves_nombres_fin[titulos_celdas_nombres])]

			celdas_ccn 		= hoja1[str(self.claves_ccn_inicio[titulos_celdas]):
									str(self.claves_ccn_fin[titulos_celdas])]



			for fila_ccn, fila_nomb in zip(celdas_ccn,celdas_nombres):
				for celda_titulo,celda_titulo_nomb in zip(fila_ccn,fila_nomb):

					if titulos_celdas_nombres.strip()   == "No. Control":
						self.no_control.append(celda_titulo_nomb.value)

					elif titulos_celdas == 793:
						self.ccn793.append(celda_titulo.value)

					elif titulos_celdas == 794:
						self.ccn794.append(celda_titulo.value)


					elif titulos_celdas == 1000:
						self.ccn1000.append(celda_titulo.value)

					elif titulos_celdas == 1101:
						self.ccn1101.append(celda_titulo.value)

					elif titulos_celdas == 1102:
						self.ccn1102.append(celda_titulo.value)

					elif titulos_celdas == 1104:
						self.ccn1104.append(celda_titulo.value)

					elif titulos_celdas == 1112:
						self.ccn1112.append(celda_titulo.value)

					elif titulos_celdas == 1114:
						self.ccn1114.append(celda_titulo.value)

					elif titulos_celdas == 1115:
						self.ccn1115.append(celda_titulo.value)

					elif titulos_celdas == 1116:
						self.ccn1116.append(celda_titulo.value)

					elif titulos_celdas == 1119:
						self.ccn1119.append(celda_titulo.value)

					elif titulos_celdas == 1129:
						self.ccn1129.append(celda_titulo.value)

					elif titulos_celdas == 1133:
						self.ccn1133.append(celda_titulo.value)

					elif titulos_celdas == 1136:
						self.ccn1136.append(celda_titulo.value)

					elif titulos_celdas == 1148:
						self.ccn1148.append(celda_titulo.value)

					elif titulos_celdas == 1159:
						self.ccn1159.append(celda_titulo.value)

					elif titulos_celdas == 1401:
						self.ccn1401.append(celda_titulo.value)

					elif titulos_celdas == 1415:
						self.ccn1415.append(celda_titulo.value)

					elif titulos_celdas == 1501:
						self.ccn1501.append(celda_titulo.value)

					elif titulos_celdas == 1502:
						self.ccn1502.append(celda_titulo.value)

					elif titulos_celdas == 1512:
						self.ccn1512.append(celda_titulo.value)

					elif titulos_celdas == 1536:
						self.ccn1536.append(celda_titulo.value)

					elif titulos_celdas == 1610:
						self.ccn1610.append(celda_titulo.value)

					elif titulos_celdas == 1620:
						self.ccn1620.append(celda_titulo.value)

					elif titulos_celdas == 1645:
						self.ccn1645.append(celda_titulo.value)

					elif titulos_celdas == 2140:
						self.ccn2140.append(celda_titulo.value)

					elif titulos_celdas == 2141:
						self.ccn2141.append(celda_titulo.value)

					elif titulos_celdas == 2144:
						self.ccn2144.append(celda_titulo.value)

					elif titulos_celdas == 2145:
						self.ccn2145.append(celda_titulo.value)

					elif titulos_celdas == 2148:
						self.ccn2148.append(celda_titulo.value)


					elif titulos_celdas == 2150:
						self.ccn2150.append(celda_titulo.value)

					elif titulos_celdas == 2151:
						self.ccn2151.append(celda_titulo.value)

					elif titulos_celdas == 2155:
						self.ccn2155.append(celda_titulo.value)

					elif titulos_celdas == 2200:
						self.ccn2200.append(celda_titulo.value)

					elif titulos_celdas == 2240:
						self.ccn2240.append(celda_titulo.value)

					elif titulos_celdas == 2450:
						self.ccn2450.append(celda_titulo.value)

					elif titulos_celdas == 2550:
						self.ccn2550.append(celda_titulo.value)

					elif titulos_celdas == 2566:
						self.ccn2566.append(celda_titulo.value)

					elif titulos_celdas == 2587:
						self.ccn2587.append(celda_titulo.value)

					elif titulos_celdas == 2596:
						self.ccn2596.append(celda_titulo.value)

					elif titulos_celdas == 2597:
						self.ccn2597.append(celda_titulo.value)

					elif titulos_celdas == 2598:
						self.ccn2598.append(celda_titulo.value)

					elif titulos_celdas == 2691:
						self.ccn2691.append(celda_titulo.value)

					elif titulos_celdas == "/300":
						self.ccn300.append(celda_titulo.value)

					elif titulos_celdas == "/305":
						self.ccn305.append(celda_titulo.value)

					elif titulos_celdas == "/306":
						self.ccn306.append(celda_titulo.value)

					elif titulos_celdas == "/380":
						self.ccn380.append(celda_titulo.value)

					elif titulos_celdas == "/391":
						self.ccn391.append(celda_titulo.value)

					elif titulos_celdas == "/481":
						self.ccn481.append(celda_titulo.value)

					elif titulos_celdas == "/494":
						self.ccn494.append(celda_titulo.value)

					elif titulos_celdas == "/559":
						self.ccn559.append(celda_titulo.value)



		self.barra_progreso.step(80)
		self.doc_iq.close()







		"""OBTENER EMPLEADOS DE BASE

			for titulos_celdas,titulos_ccn in zip(self.celdas_dicc_nombres,self.celdas_dicc_ccn):
				celdas_conte 		 = hoja1[str(self.claves_nombres_inicio["Área de Personal"]):str(self.claves_nombres_fin["Área de Personal"])]
				celdas_conte_control = hoja1[str(self.claves_nombres_inicio["No. Control"]):str(self.claves_nombres_fin["No. Control"])]
				celdas_conte_ccn	 = hoja1[str(self.claves_ccn_inicio[titulos_ccn]):str(self.claves_ccn_fin[titulos_ccn])]

				for fila,fila_control,fila_ccn in zip(celdas_conte,celdas_conte_control,celdas_conte_ccn):
					for celda,celda_control,celda_ccn in zip(fila,fila_control,fila_ccn):
						if celda.value == "BASE SINDICALIZADO" or celda.value == "BASE NO SINDICALIZAD":
							self.empleados_dicc[celda_control.value]={titulos_ccn:{celda_ccn.value}
							self.empleados_list.append(celda_control.value)

		for celdas in self.celdas_dicc_nombres:

			if celdas == "Área de Personal":

				celdas_conte = hoja1[str(self.claves_nombres_inicio["Área de Personal"]):str(self.claves_nombres_fin["Área de Personal"])]

				for fila in celdas_conte:
					for celda in fila:
						#print(celda.value)
						if celda.value == "BASE SINDICALIZADO" or celda.value == "BASE NO SINDICALIZAD":
							#print("celda_area de Personal.... ",celda.value)

							for celdas in self.celdas_dicc_ccn:

								celdas_por_copiar = hoja1[str(self.claves_ccn_inicio[celdas]):str(self.claves_ccn_fin[celdas])]

								for fila in celdas_por_copiar:
									for celda in fila:
										#self.i.append(celda.value)
										self.conte_iq[celdas]= celda.value"""


	def escritura_en_excel(self):
		"""Escribe toda la info recuperada de los archivos
			leidos y la escribe en el archivo de IQ_retimbre"""


		self.lectura_IQ()
		self.lectura_reporte_SAP()
		self.control_de_listados()
		self.lectura_de_UUID()



		self.ruta_excel_trabajo   			= self.documentos_dicc["IQ_TRABAJO"]
		self.wb        						= openpyxl.load_workbook(self.ruta_excel_trabajo)
		self.wb.active						= 0





		"""REPORTE IQ """

		for no_control in self.no_control:
			self.wb.active= 0
			self.control_fila_iq +=1


			self.c_Control    =	self.wb.active.cell(row = int(self.control_fila_iq),column = 4)
			self.c_Control.value = (no_control)





		if (self.ccn1401 ==  True or self.ccn2240 ==  True or
			self.ccn2566 ==  True or self.ccn481 == True or
			self.ccn559 ==  True):


			for ccn2240,ccn481,ccn1401, in zip(self.ccn2240,self.ccn481,self.ccn1401):
				self.wb.active= 0
				self.ccn1401_fila_iq  +=1
				self.ccn2240_fila_iq  +=1
				self.ccn481_fila_iq  +=1




				self.c_1401    =	self.wb.active.cell(row = int(self.ccn1401_fila_iq),column = 20)
				self.c_1401.value = (ccn1401)

				self.c_2240    =	self.wb.active.cell(row = int(self.ccn2240_fila_iq),column = 21)
				self.c_2240.value = (ccn2240)


				self.c_481    =	self.wb.active.cell(row = int(self.ccn481_fila_iq),column = 23)
				self.c_481.value = (ccn481)

			for ccn559,ccn2566 in zip(self.ccn559,self.ccn2566):

				self.wb.active= 0
				self.ccn559_fila_iq  +=1
				self.ccn2566_fila_iq +=1

				self.c_599    =	self.wb.active.cell(row = int(self.ccn559_fila_iq),column = 24)
				self.c_599.value = (ccn559)

				self.c_2566    =	self.wb.active.cell(row = int(self.ccn2566_fila_iq),column = 22)
				self.c_2566.value = (ccn2566)

			for titulos in self.titulos_completos_iq:

				if titulos.strip() == "1401 APORTACION SEGURIDAD S.":

					self.t_1401    =	self.wb.active.cell(row = 1, column = 20)
					self.t_1401.value = (titulos)


				elif titulos.strip() == "2240 IMPUESTO ORDINARIO":
					self.t_2240    =	self.wb.active.cell(row = 1, column = 21)
					self.t_2240.value = (titulos)


				elif titulos.strip() == "2566 IMPUESTO EXTRAORDINARIO":
					self.t_2566    =	self.wb.active.cell(row = 1, column = 22)
					self.t_2566.value = (titulos)


				elif titulos.strip() == "/481 Subsidio al empleo efvo":
					self.t_481    =	self.wb.active.cell(row = 1, column = 23)
					self.t_481.value = (titulos)



				elif titulos.strip() == "/559 Transferencia bancaria":
					self.t_559    =	self.wb.active.cell(row = 1, column = 24)
					self.t_559.value = (titulos)

		else:

			if self.ccn481 == False and self.ccn2566 == False:

				for ccn2240,ccn1401 in zip(float(self.ccn2240), float(self.ccn1401)):
					self.wb.active= 0
					self.ccn1401_fila_iq  +=1
					self.ccn2240_fila_iq  +=1





					self.c_1401    =	self.wb.active.cell(row = int(self.ccn1401_fila_iq),column = 20)
					self.c_1401.value = (ccn1401)

					self.c_2240    =	self.wb.active.cell(row = int(self.ccn2240_fila_iq),column = 21)
					self.c_2240.value = (ccn2240)


				for ccn559 in self.ccn559:

					self.wb.active= 0
					self.ccn559_fila_iq  +=1


					self.c_599    =	self.wb.active.cell(row = int(self.ccn559_fila_iq),column = 22)
					self.c_599.value = (ccn559)


				for titulos in self.titulos_completos_iq:

					if titulos.strip() == "1401 APORTACION SEGURIDAD S.":

						self.t_1401    =	self.wb.active.cell(row = 1, column = 20)
						self.t_1401.value = (titulos)


					elif titulos.strip() == "2240 IMPUESTO ORDINARIO":
						self.t_2240    =	self.wb.active.cell(row = 1, column = 21)
						self.t_2240.value = (titulos)

					elif titulos.strip() == "/559 Transferencia bancaria":
						self.t_559    =	self.wb.active.cell(row = 1, column = 22)
						self.t_559.value = (titulos)



			elif self.ccn481 == False:

				for ccn2240,ccn1401 in zip(self.ccn2240, self.ccn1401):
					self.wb.active= 0
					self.ccn1401_fila_iq  +=1
					self.ccn2240_fila_iq  +=1





					self.c_1401    =	self.wb.active.cell(row = int(self.ccn1401_fila_iq),column = 20)
					self.c_1401.value = (ccn1401)

					self.c_2240    =	self.wb.active.cell(row = int(self.ccn2240_fila_iq),column = 21)
					self.c_2240.value = (ccn2240)


				for ccn559,ccn2566 in zip(self.ccn559,self.ccn2566):

					self.wb.active= 0
					self.ccn559_fila_iq  +=1
					self.ccn2566_fila_iq +=1

					self.c_599    =	self.wb.active.cell(row = int(self.ccn559_fila_iq),column = 23)
					self.c_599.value = (ccn559)

					self.c_2566    =	self.wb.active.cell(row = int(self.ccn2566_fila_iq),column = 22)
					self.c_2566.value = (ccn2566)

				for titulos in self.titulos_completos_iq:

					if titulos.strip() == "1401 APORTACION SEGURIDAD S.":

						self.t_1401    =	self.wb.active.cell(row = 1, column = 20)
						self.t_1401.value = (titulos)


					elif titulos.strip() == "2240 IMPUESTO ORDINARIO":
						self.t_2240    =	self.wb.active.cell(row = 1, column = 21)
						self.t_2240.value = (titulos)


					elif titulos.strip() == "2566 IMPUESTO EXTRAORDINARIO":
						self.t_2566    =	self.wb.active.cell(row = 1, column = 22)
						self.t_2566.value = (titulos)

					elif titulos.strip() == "/559 Transferencia bancaria":
						self.t_559    =	self.wb.active.cell(row = 1, column = 23)
						self.t_559.value = (titulos)


			elif self.ccn2566 == False:


				for ccn2240,ccn481,ccn1401 in zip(self.ccn2240,self.ccn481,self.ccn1401):

					self.ccn1401_fila_iq  +=1
					self.ccn2240_fila_iq  +=1
					self.ccn481_fila_iq  +=1

					self.wb.active	= 0


					self.c_1401    =	self.wb.active.cell(row = int(self.ccn1401_fila_iq),column = 20)
					self.c_1401.value = (ccn1401)

					self.c_2240    =	self.wb.active.cell(row = int(self.ccn1401_fila_iq),column = 21)
					self.c_2240.value = (ccn2240)


					self.c_481    =	self.wb.active.cell(row = int(self.ccn481_fila_iq),column = 22)
					self.c_481.value = (ccn481)

				for ccn559 in self.ccn559:
					self.wb.active= 0
					self.ccn559_fila_iq  +=1

					self.c_599    =	self.wb.active.cell(row = int(self.ccn559_fila_iq),column = 23)
					self.c_599.value = (ccn559)


				for titulos in self.titulos_completos_iq:

					if titulos.strip() == "1401 APORTACION SEGURIDAD S.":

						self.t_1401    =	self.wb.active.cell(row = 1, column = 20)
						self.t_1401.value = (titulos)

					elif titulos.strip() == "2240 IMPUESTO ORDINARIO":
						self.t_2240    =	self.wb.active.cell(row = 1, column = 21)
						self.t_2240.value = (titulos)


					elif titulos.strip() == "/481 Subsidio al empleo efvo":
						self.t_481    =	self.wb.active.cell(row = 1, column = 22)
						self.t_481.value = (titulos)


					elif titulos.strip() == "/559 Transferencia bancaria":
						self.t_559    =	self.wb.active.cell(row = 1, column = 23)
						self.t_559.value = (titulos)














		"""REPORTE SAP_TXT"""
		for control_sap,nom1_sap,nom4_sap in zip(self.control, self.nom1, self.nom4):
			#print(control_sap," ",nom1_sap," ", nom4_sap )

			self.fila_control_sap +=1
			self.fila_nom1_sap	  +=1
			self.fila_nom4_sap	  +=1


			self.wb.active	= 3

			self.c_Control_sap     =	self.wb.active.cell(row = int(self.fila_control_sap),
															column = int(self.columna_control_sap))
			self.c_Control_sap.value = (control_sap)

			self.c_nom1_sap      = self.wb.active.cell(row = (self.fila_nom1_sap),
														 column = (self.columna_nom1_sap))
			self.c_nom1_sap.value = (nom1_sap)

			self.c_nom4_sap      = self.wb.active.cell(row = (self.fila_nom4_sap),
														 column = (self.columna_nom4_sap))
			self.c_nom4_sap.value = (nom4_sap)


			self.c_control_titulo           = self.wb.active.cell(row = 1, column = 1)
			self.c_control_titulo.value     = ("NO CONTROL")

			self.c_nom1_titulo          	= self.wb.active.cell(row = 1, column = 2)
			self.c_nom1_titulo.value  	    = ("NOM1")

			self.c_nom4_titulo          	= self.wb.active.cell(row = 1, column = 3)
			self.c_nom4_titulo.value   	    = ("NOM4")



		"""RUTAS_ARCHIVOS TXT Y XML"""
		for control_txt1,ruta_archivos_txt1 in zip(self.control_list_txt1,
												   self.ruta_completa_archivos_list_txt1):
			#print(control_txt1 + "  " + ruta_archivos_txt1)

			self.control_fila_txt1 +=1
			self.path_fila_txt1    +=1

			self.wb.active	= 4

			self.c_Control_text       =	self.wb.active.cell(row = int(self.control_fila_txt1),
															column = int(self.control_columna_txt1))
			self.c_Control_text.value = (int(control_txt1))

			self.c_path_text       = self.wb.active.cell(row = (self.path_fila_txt1),
														 column = (self.path_columna_txt1))
			self.c_path_text.value = (ruta_archivos_txt1)


			self.c_Control_titulo       = self.wb.active.cell(row = int(self.titulo_fila_control),
															  column = int(self.titulo_columna_control))
			self.c_Control_titulo.value = ("CONTROL")

			self.c_path_titulo          = self.wb.active.cell(row = (self.titulo_fila_path),
															  column = (self.titulo_columna_path))
			self.c_path_titulo .value   = ("PATH")




		for control_txt4,ruta_archivos_txt4 in zip(self.control_list_txt4,
												   self.ruta_completa_archivos_list_txt4):


			self.control_fila_txt4 +=1
			self.path_fila_txt4 	  +=1

			self.wb.active	= 5
			self.c_Control_text       = self.wb.active.cell(row = int(self.control_fila_txt4),
															column = int(self.control_columna_txt4))
			self.c_Control_text.value = (int(control_txt4))

			self.c_path_text       = self.wb.active.cell(row = (self.path_fila_txt4),
													     column = (self.path_columna_txt4))
			self.c_path_text.value = (ruta_archivos_txt4)


			self.c_Control_titulo       = self.wb.active.cell(row = int(self.titulo_fila_control),
															  column = int(self.titulo_columna_control))
			self.c_Control_titulo.value = ("CONTROL")

			self.c_path_titulo          = self.wb.active.cell(row = (self.titulo_fila_path),
															  column = (self.titulo_columna_path))
			self.c_path_titulo .value   = ("PATH")



		for control_xml1,ruta_archivos_xml1,uuid_xml1,archiv_xml1 in zip(self.control_list_xml1,
																		 self.ruta_completa_archivos_list_xml1,
																		 self.uuid_xml1_lista, self.archivo_xml1_lista):
			#print(control_xml1 + " " + ruta_archivos_xml1)
			self.control_fila_xml1 +=1
			self.path_fila_xml1	  +=1
			self.fila_xml1_uuid +=1
			self.fila_xml1_arch +=1



			self.wb.active	= 6
			self.c_Control_text       =	self.wb.active.cell(row = int(self.control_fila_xml1),
															column = int(self.control_columna_xml1))
			self.c_Control_text.value = (int(control_xml1))

			self.c_path_text       	  = self.wb.active.cell(row = int(self.path_fila_xml1),
															column = int(self.path_columna_xml1))
			self.c_path_text.value    = (ruta_archivos_xml1)

			self.c_uuid_text       	  = self.wb.active.cell(row = int(self.fila_xml1_uuid),
														    column = int(self.columna_xml1_uuid))
			self.c_uuid_text.value    = (uuid_xml1)

			#self.c_arch_text       	  = self.wb.active.cell(row = int(self.fila_xml1_arch),
																#column = int(self.columna_xml1_arch))
			#elf.c_arch_text.value    = (archiv_xml1)





			self.c_Control_titulo       = self.wb.active.cell(row = int(self.titulo_fila_control),
															  column = int(self.titulo_columna_control))
			self.c_Control_titulo.value = ("CONTROL")

			self.c_path_titulo          = self.wb.active.cell(row = (self.titulo_fila_path),
															  column = (self.titulo_columna_path))
			self.c_path_titulo.value   = ("PATH")


			self.c_uuid_titulo          = self.wb.active.cell(row = 1, column = 3)
			self.c_uuid_titulo.value   = ("UUID")
			#self.c_uuid_titulo          = self.wb.active.cell(row = 1, column = 4)
			#self.c_uuid_titulo.value   = ("ARCHIVO")


		for control_xml4,ruta_archivos_xml4,uuid_xml4, archivo_xml4 in zip(self.control_list_xml4,
																		   self.ruta_completa_archivos_list_xml4,
												   						   self.uuid_xml4_lista, self.archivo_xml4_lista):
			#print(control_xml4 + "  " + ruta_archivos_xml4)
			self.control_fila_xml4 +=1
			self.path_fila_xml4	  +=1
			self.wb.active	= 7
			self.fila_xml4_uuid	+=1
			self.fila_xml4_arch +=1


			self.c_Control_text       = self.wb.active.cell(row = int(self.control_fila_xml4),
															column = int(self.control_columna_xml4))
			self.c_Control_text.value = (int(control_xml4))

			self.c_path_text       	  = self.wb.active.cell(row = (self.path_fila_xml4),
															column = (self.path_columna_xml4))
			self.c_path_text.value    = (ruta_archivos_xml4)


			self.c_uuid_text       	  = self.wb.active.cell(row = (self.fila_xml4_uuid),
															column = (self.columna_xml14_uuid))
			self.c_uuid_text.value    = (uuid_xml4)

			#self.c_archivo_text       	  = self.wb.active.cell(row = (self.fila_xml4_arch),
															#column = (self.columna_xml4_arch))
			#elf.c_archivo_text.value    = (archivo_xml4)




			self.c_Control_titulo       = self.wb.active.cell(row = int(self.titulo_fila_control),
																  column = int(self.titulo_columna_control))
			self.c_Control_titulo.value = ("CONTROL")

			self.c_path_titulo          = self.wb.active.cell(row = (self.titulo_fila_path),
																  column = (self.titulo_columna_path))
			self.c_path_titulo .value   = ("PATH")

			self.c_uuid_titulo          = self.wb.active.cell(row = 1,column =3)
			self.c_uuid_titulo .value   = ("UUID")

			#self.c_archivo_titulo          = self.wb.active.cell(row = 1,column =4)
			#self.c_archivo_titulo .value   = ("ARCHIVO")



		self.path_save = asksaveasfile()
		self.wb.save(self.path_save.name)
		self.salida_texto.insert(INSERT, "Se a guardado el documento en: '\n' {}".format(self.path_save.name))
		self.barra_progreso.step(100)


	def guardar_excel(self):


		self.escritura_en_excel()



	def mostrar_interfaz(self):
		"""Muestra ka interfaz Grafica del Programa"""

		#BOTON ABRIR DOCUMENTOS DE TRABAJO
		self.boton_cargar_docs=tk.Button(self.ventana_prin, text="ABRIR_DOCS",
										 command=carpetas.rutas_docs,
										 background="#727272")
		self.boton_cargar_docs.place(x=10, y=20)

		#BOTON PARA LEER ARCHIVOS
		self.boton_leer=tk.Button(self.ventana_prin,
								  text="LEER SUBCARPETAS",
								  command=carpetas.lectura_de_carpeta,
								  background="#727272")

		self.boton_leer.place(x=100, y=20)

		#BOTON 	STAR
		self.boton_empezar=tk.Button(self.ventana_prin,
									 text="Empezar",
									 command=carpetas.escritura_en_excel,
									 background="#727272")

		self.boton_empezar.place(x=225, y=20)

		"""self.boton_guardar=tk.Button(self.ventana_prin,
															 text="GUARDAR RETIMBRE",
															 command=carpetas.guardar_excel,
															 background="#727272")

								self.boton_guardar.place(x=300, y=20)"""




		self.salida_texto.place(x=30,y=150)

		self.barra_progreso.place(x=30,y=400, width=720)
		#mostramos la  ventana
		self.ventana_prin.mainloop()



carpetas = Carpetas()
carpetas.mostrar_interfaz()
